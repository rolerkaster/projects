import os
from datetime import datetime, timedelta
from openpyxl import Workbook
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.application import MIMEApplication
import json
from app.utils.database import get_db_connection
from app.utils.logger import setup_logger

# Создаем директории если их нет
os.makedirs('logs', exist_ok=True)
os.makedirs('reports', exist_ok=True)

# Настраиваем логгер
logger = setup_logger('ReportController', 'reports.log')

class ReportController:
    def __init__(self):
        self.report_interval = int(os.getenv('REPORT_TIME_INTERVAL', 24))
        self.admin_emails = os.getenv('ADMIN_EMAILS', '').split(',')
        self.report_format = os.getenv('REPORT_FORMAT', 'json')
        self.reports_dir = 'reports'
        self.keep_reports = True  # Флаг для сохранения отчетов

    def generate_report(self):
        """Генерация отчета о статистике использования системы"""
        try:
            logger.info("Starting report generation")
            
            # Получаем временной интервал
            end_date = datetime.utcnow()
            start_date = end_date - timedelta(hours=self.report_interval)
            
            # Собираем данные
            method_stats = self._get_method_stats(start_date, end_date)
            entity_stats = self._get_entity_stats(start_date, end_date)
            user_stats = self._get_user_stats(start_date, end_date)
            
            # Формируем отчет
            report_path = self._create_report(method_stats, entity_stats, user_stats, start_date, end_date)
            
            # Отправляем отчет
            self._send_report(report_path)
            
            # Сохраняем отчет если нужно
            if not self.keep_reports:
                os.remove(report_path)
                logger.info(f"Report {report_path} deleted")
            else:
                logger.info(f"Report saved at {report_path}")
            
            logger.info("Report generation completed successfully")
            return True
        except Exception as e:
            logger.error(f"Error generating report: {str(e)}")
            return False

    def _get_method_stats(self, start_date, end_date):
        """Получение статистики по методам"""
        conn = get_db_connection()
        cursor = conn.cursor()
        
        cursor.execute("""
            SELECT 
                http_method || ' ' || api_path as method,
                COUNT(*) as call_count,
                MAX(created_at) as last_call
            FROM LogsRequests
            WHERE created_at BETWEEN ? AND ?
            GROUP BY http_method, api_path
            ORDER BY call_count DESC
        """, (start_date.isoformat(), end_date.isoformat()))
        
        result = []
        for row in cursor.fetchall():
            result.append({
                'method': row[0],
                'call_count': row[1],
                'last_call': row[2]
            })
        
        conn.close()
        return result

    def _get_entity_stats(self, start_date, end_date):
        """Получение статистики по редактируемым сущностям"""
        conn = get_db_connection()
        cursor = conn.cursor()
        
        cursor.execute("""
            SELECT 
                CASE 
                    WHEN api_path LIKE '%/user%' THEN 'User'
                    WHEN api_path LIKE '%/role%' THEN 'Role'
                    WHEN api_path LIKE '%/permission%' THEN 'Permission'
                    ELSE 'Other'
                END as entity_type,
                COUNT(*) as change_count,
                MAX(created_at) as last_change
            FROM LogsRequests
            WHERE created_at BETWEEN ? AND ?
                AND http_method IN ('POST', 'PUT', 'DELETE')
            GROUP BY 
                CASE 
                    WHEN api_path LIKE '%/user%' THEN 'User'
                    WHEN api_path LIKE '%/role%' THEN 'Role'
                    WHEN api_path LIKE '%/permission%' THEN 'Permission'
                    ELSE 'Other'
                END
            ORDER BY change_count DESC
        """, (start_date.isoformat(), end_date.isoformat()))
        
        result = []
        for row in cursor.fetchall():
            result.append({
                'entity_type': row[0],
                'change_count': row[1],
                'last_change': row[2]
            })
        
        conn.close()
        return result

    def _get_user_stats(self, start_date, end_date):
        """Получение статистики по пользователям"""
        conn = get_db_connection()
        cursor = conn.cursor()
        
        cursor.execute("""
            SELECT 
                user_id,
                COUNT(*) as request_count,
                COUNT(CASE WHEN api_path LIKE '%/auth%' THEN 1 END) as auth_count,
                COUNT(CASE WHEN http_method IN ('POST', 'PUT', 'DELETE') THEN 1 END) as change_count,
                MAX(created_at) as last_activity
            FROM LogsRequests
            WHERE created_at BETWEEN ? AND ?
                AND user_id IS NOT NULL
            GROUP BY user_id
            ORDER BY request_count DESC
        """, (start_date.isoformat(), end_date.isoformat()))
        
        result = []
        for row in cursor.fetchall():
            result.append({
                'user_id': row[0],
                'request_count': row[1],
                'auth_count': row[2],
                'change_count': row[3],
                'last_activity': row[4]
            })
        
        conn.close()
        return result

    def _create_report(self, method_stats, entity_stats, user_stats, start_date, end_date):
        """Создание отчета в выбранном формате"""
        if self.report_format == 'xlsx':
            return self._create_excel_report(method_stats, entity_stats, user_stats, start_date, end_date)
        elif self.report_format == 'json':
            return self._create_json_report(method_stats, entity_stats, user_stats, start_date, end_date)
        else:
            raise ValueError(f"Unsupported report format: {self.report_format}")

    def _get_unique_filename(self, base_filename):
        """Получение уникального имени файла"""
        counter = 1
        filename = base_filename
        while os.path.exists(filename):
            name, ext = os.path.splitext(base_filename)
            filename = f"{name}_{counter}{ext}"
            counter += 1
        return filename

    def _create_excel_report(self, method_stats, entity_stats, user_stats, start_date, end_date):
        """Создание отчета в формате Excel"""
        wb = Workbook()
        
        # Статистика методов
        ws = wb.active
        ws.title = "Method Statistics"
        ws.append(["Method", "Call Count", "Last Call"])
        for stat in method_stats:
            ws.append([stat['method'], stat['call_count'], stat['last_call']])
        
        # Статистика сущностей
        ws = wb.create_sheet("Entity Statistics")
        ws.append(["Entity", "Change Count", "Last Change"])
        for stat in entity_stats:
            ws.append([stat['entity_type'], stat['change_count'], stat['last_change']])
        
        # Статистика пользователей
        ws = wb.create_sheet("User Statistics")
        ws.append(["User", "Request Count", "Auth Count", "Change Count", "Last Activity"])
        for stat in user_stats:
            ws.append([
                stat['user_id'],
                stat['request_count'],
                stat['auth_count'],
                stat['change_count'],
                stat['last_activity']
            ])
        
        # Информация об отчете
        ws = wb.create_sheet("Report Info")
        ws.append(["Report Type", "System Usage Statistics"])
        ws.append(["Start Date", start_date.isoformat()])
        ws.append(["End Date", end_date.isoformat()])
        ws.append(["Generated At", datetime.utcnow().isoformat()])
        
        # Сохраняем файл в директорию reports с уникальным именем
        base_filename = os.path.join(self.reports_dir, f"report_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx")
        filename = self._get_unique_filename(base_filename)
        wb.save(filename)
        return filename

    def _create_json_report(self, method_stats, entity_stats, user_stats, start_date, end_date):
        """Создание отчета в формате JSON"""
        report = {
            "report_info": {
                "type": "System Usage Statistics",
                "start_date": start_date.isoformat(),
                "end_date": end_date.isoformat(),
                "generated_at": datetime.utcnow().isoformat()
            },
            "method_statistics": method_stats,
            "entity_statistics": entity_stats,
            "user_statistics": user_stats
        }
        
        # Сохраняем файл в директорию reports с уникальным именем
        base_filename = os.path.join(self.reports_dir, f"report_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.json")
        filename = self._get_unique_filename(base_filename)
        with open(filename, 'w', encoding='utf-8') as f:
            json.dump(report, f, indent=2, ensure_ascii=False)
        return filename

    def _send_report(self, report_path):
        """Отправка отчета администраторам по email"""
        try:
            # Настройки SMTP
            smtp_server = os.getenv('SMTP_SERVER')
            smtp_port = int(os.getenv('SMTP_PORT', 587))
            smtp_username = os.getenv('SMTP_USERNAME')
            smtp_password = os.getenv('SMTP_PASSWORD')
            
            # Создаем сообщение
            msg = MIMEMultipart()
            msg['Subject'] = f'System Usage Report - {datetime.utcnow().strftime("%Y-%m-%d")}'
            msg['From'] = smtp_username
            msg['To'] = ', '.join(self.admin_emails)
            
            # Добавляем текст
            body = "Please find attached the system usage report."
            msg.attach(MIMEText(body, 'plain'))
            
            # Добавляем файл отчета
            with open(report_path, 'rb') as f:
                part = MIMEApplication(f.read(), Name=os.path.basename(report_path))
                part['Content-Disposition'] = f'attachment; filename="{os.path.basename(report_path)}"'
                msg.attach(part)
            
            # Отправляем
            with smtplib.SMTP(smtp_server, smtp_port) as server:
                server.starttls()
                server.login(smtp_username, smtp_password)
                server.send_message(msg)
            
            logger.info(f"Report sent successfully to {', '.join(self.admin_emails)}")
        except Exception as e:
            logger.error(f"Error sending report: {str(e)}")
            raise 