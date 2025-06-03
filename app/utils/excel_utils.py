from typing import List, Dict, Any
from openpyxl import Workbook, load_workbook
import sqlite3
import logging
from openpyxl.utils.exceptions import InvalidFileException
import secrets
import hashlib
import string
from app.utils.validation import validate_username, validate_email, validate_birthday, validate_password
from flask import request
import jwt
import os

def get_current_user():
    """Получает имя текущего пользователя из JWT токена."""
    try:
        auth_header = request.headers.get('Authorization')
        if not auth_header or not auth_header.startswith('Bearer '):
            return 'system'
            
        token = auth_header.split(' ')[1]
        secret_key = os.getenv('JWT_SECRET_KEY', 'your-secret-key')
        payload = jwt.decode(token, secret_key, algorithms=['HS256'])
        return payload.get('username', 'system')
    except:
        return 'system'

def get_db_connection():
    conn = sqlite3.connect('database.db')
    conn.row_factory = sqlite3.Row
    return conn

def generate_temp_password():
    """Генерирует временный пароль и его хеш."""
    # Генерируем пароль, соответствующий требованиям валидации
    lowercase = string.ascii_lowercase
    uppercase = string.ascii_uppercase
    digits = string.digits
    special = "!@#$%^&*(),.?\":{}|<>"
    
    # Гарантируем наличие всех требуемых типов символов
    password = [
        secrets.choice(uppercase),  # Одна заглавная буква
        secrets.choice(lowercase),  # Одна строчная буква
        secrets.choice(digits),     # Одна цифра
        secrets.choice(special),    # Один специальный символ
    ]
    
    # Добавляем еще 4 случайных символов для достижения минимальной длины
    all_chars = lowercase + uppercase + digits + special
    password.extend(secrets.choice(all_chars) for _ in range(4))
    
    # Перемешиваем символы
    password = list(password)
    secrets.SystemRandom().shuffle(password)
    temp_password = ''.join(password)
    
    # Проверяем, что пароль соответствует требованиям
    if not validate_password(temp_password):
        # Если вдруг не соответствует, генерируем заново
        return generate_temp_password()
        
    logging.info(f"Generated valid temporary password")
    
    # Используем тот же метод хеширования, что и в auth_controller
    password_hash = hashlib.sha256(temp_password.encode()).hexdigest()
    return temp_password, password_hash

class ExcelExporter:
    @staticmethod
    def export_entities(entity_type: str, exclude_columns: List[str] = None) -> Workbook:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # Получаем все записи из таблицы
        if entity_type == 'users':
            cursor.execute('SELECT username, email, birthday FROM Users')
        elif entity_type == 'roles':
            cursor.execute('SELECT name, description, code FROM Roles WHERE deleted_at IS NULL')
        else:
            raise ValueError("Неподдерживаемый тип сущности")
            
        records = cursor.fetchall()
        
        if not records:
            conn.close()
            return Workbook()

        wb = Workbook()
        ws = wb.active
        
        # Получаем заголовки колонок
        columns = [description[0] for description in cursor.description 
                  if not exclude_columns or description[0] not in exclude_columns]
        
        # Записываем заголовки
        for col, header in enumerate(columns, 1):
            ws.cell(row=1, column=col, value=header)
        
        # Записываем данные
        for row, record in enumerate(records, 2):
            for col, column in enumerate(columns, 1):
                ws.cell(row=row, column=col, value=record[column])
        
        conn.close()
        return wb

class ImportResult:
    def __init__(self):
        self.success_added: List[Dict[str, Any]] = []
        self.success_updated: List[Dict[str, Any]] = []
        self.errors: List[Dict[str, Any]] = []
        self.duplicates: List[Dict[str, Any]] = []

class ImportMode:
    ADD_ONLY = "add_only"
    ADD_OR_UPDATE = "add_or_update"

class ErrorHandlingMode:
    CONTINUE_ON_ERROR = "continue_on_error"
    STOP_ON_ERROR = "stop_on_error"
    ROLLBACK_ON_ERROR = "rollback_on_error"

class ExcelImporter:
    @staticmethod
    def import_entities(
        file_path: str,
        entity_type: str,
        import_mode: str = ImportMode.ADD_ONLY,
        error_handling: str = ErrorHandlingMode.CONTINUE_ON_ERROR,
        exclude_columns: List[str] = None
    ) -> ImportResult:
        result = ImportResult()
        conn = get_db_connection()
        cursor = conn.cursor()
        
        try:
            logging.info(f"Attempting to load workbook from {file_path}")
            try:
                wb = load_workbook(filename=file_path, read_only=True, data_only=True)
            except InvalidFileException as e:
                logging.error(f"Invalid Excel file: {str(e)}")
                raise ValueError("Неверный формат Excel файла")
            except Exception as e:
                logging.error(f"Error loading Excel file: {str(e)}")
                raise ValueError("Ошибка при чтении Excel файла")
                
            ws = wb.active
            if not ws:
                raise ValueError("Excel файл не содержит листов")

            # Получаем заголовки
            headers = []
            for row in ws.iter_rows(min_row=1, max_row=1):
                headers = [cell.value for cell in row if cell.value is not None]
                break
                
            if not headers:
                raise ValueError("Excel файл не содержит заголовков")
                
            logging.info(f"Found headers: {headers}")
        
            # Проверяем корректность заголовков
            if entity_type == 'users':
                valid_columns = {'username', 'email', 'birthday'}
            elif entity_type == 'roles':
                valid_columns = {'name', 'description', 'code'}
            else:
                raise ValueError("Неподдерживаемый тип сущности")
                
            if not all(header in valid_columns for header in headers):
                invalid_headers = [h for h in headers if h not in valid_columns]
                raise ValueError(f"Некорректные заголовки в файле: {invalid_headers}")

            records_to_process = []
            
            # Читаем все строки
            row_count = 0
            for row_num, row in enumerate(ws.iter_rows(min_row=2), 2):
                row_count += 1
                record = {}
                empty_row = True
                
                for header, cell in zip(headers, row):
                    if not exclude_columns or header not in exclude_columns:
                        value = cell.value if cell else None
                        record[header] = value
                        if value is not None:  # Проверяем, что значение не None
                            empty_row = False
                            
                logging.info(f"Row {row_num}: {record}")
                
                if not empty_row:  # Если строка не пустая
                    # Валидация данных
                    if entity_type == 'users':
                        validation_errors = []
                        
                        if not record.get('username'):
                            validation_errors.append("Имя пользователя не может быть пустым")
                        elif not validate_username(record.get('username')):
                            validation_errors.append("Неверный формат имени пользователя. Требования: начинается с заглавной буквы, минимум 7 символов, только латинские буквы")
                            
                        if not record.get('email'):
                            validation_errors.append("Email не может быть пустым")
                        elif not validate_email(record.get('email')):
                            validation_errors.append("Неверный формат email")
                            
                        if not record.get('birthday'):
                            validation_errors.append("Дата рождения не может быть пустой")
                        elif not validate_birthday(record.get('birthday')):
                            validation_errors.append("Неверный формат даты рождения или возраст меньше 14 лет")
                            
                        if validation_errors:
                            result.errors.append({
                                "row": row_num,
                                "error": "; ".join(validation_errors)
                            })
                            continue
                            
                    elif entity_type == 'roles':
                        validation_errors = []
                        
                        if not record.get('name'):
                            validation_errors.append("Название роли не может быть пустым")
                        if not record.get('code'):
                            validation_errors.append("Код роли не может быть пустым")
                            
                        if validation_errors:
                            result.errors.append({
                                "row": row_num,
                                "error": "; ".join(validation_errors)
                            })
                            continue
                            
                    records_to_process.append((row_num, record))
                    
            logging.info(f"Found {len(records_to_process)} valid rows to process out of {row_count} total rows")

            if not records_to_process:
                raise ValueError("Excel файл не содержит валидных данных для импорта. Проверьте формат данных и требования к полям.")

            # Получаем текущего пользователя для created_by
            current_user = get_current_user()
            logging.info(f"Current user for import: {current_user}")

            try:
                for row_num, record in records_to_process:
                    try:
                        # Проверяем существование записи
                        if entity_type == 'users':
                            cursor.execute('SELECT username FROM Users WHERE username = ? OR email = ?',
                                        (record.get('username'), record.get('email')))
                        elif entity_type == 'roles':
                            cursor.execute('SELECT name FROM Roles WHERE name = ? OR code = ?',
                                        (record.get('name'), record.get('code')))
                            
                        existing_record = cursor.fetchone()

                        if existing_record:
                            if import_mode == ImportMode.ADD_ONLY:
                                result.duplicates.append({
                                    "row": row_num,
                                    "db_id": existing_record[0],
                                    "property": "username" if entity_type == 'users' else 'name'
                                })
                                continue
                            
                            # Обновляем существующую запись
                            if entity_type == 'users':
                                cursor.execute('''
                                    UPDATE Users 
                                    SET email = ?, birthday = ?
                                    WHERE username = ?
                                ''', (record['email'], record['birthday'], record['username']))
                            elif entity_type == 'roles':
                                cursor.execute('''
                                    UPDATE Roles 
                                    SET description = ?
                                    WHERE name = ?
                                ''', (record.get('description'), record['name']))
                                
                            result.success_updated.append({
                                "row": row_num,
                                "db_id": existing_record[0]
                            })
                        else:
                            # Создаем новую запись
                            if entity_type == 'users':
                                # Генерируем временный пароль для нового пользователя
                                temp_password, password_hash = generate_temp_password()
                                cursor.execute('''
                                    INSERT INTO Users (username, email, birthday, password_hash)
                                    VALUES (?, ?, ?, ?)
                                ''', (record['username'], record['email'], record['birthday'], password_hash))
                                
                                # Получаем роль User и привязываем к пользователю
                                cursor.execute('SELECT id FROM Roles WHERE code = ?', ('user',))
                                user_role = cursor.fetchone()
                                if user_role:
                                    cursor.execute('''
                                        INSERT INTO UsersAndRoles (user_id, role_id, created_at, created_by)
                                        VALUES (?, ?, CURRENT_TIMESTAMP, ?)
                                    ''', (record['username'], user_role['id'], 'system'))
                                
                                result.success_added.append({
                                    "row": row_num,
                                    "db_id": cursor.lastrowid,
                                    "temp_password": temp_password  # Добавляем временный пароль в результат
                                })
                            elif entity_type == 'roles':
                                cursor.execute('''
                                    INSERT INTO Roles (name, description, code, created_by)
                                    VALUES (?, ?, ?, ?)
                                ''', (record['name'], record.get('description'), record['code'], current_user))
                                
                                result.success_added.append({
                                    "row": row_num,
                                    "db_id": cursor.lastrowid
                                })

                    except Exception as e:
                        logging.error(f"Error processing row {row_num}: {str(e)}")
                        result.errors.append({
                            "row": row_num,
                            "error": str(e)
                        })
                        
                        if error_handling == ErrorHandlingMode.STOP_ON_ERROR:
                            raise e

                if error_handling == ErrorHandlingMode.ROLLBACK_ON_ERROR and result.errors:
                    conn.rollback()
                    result.success_added.clear()
                    result.success_updated.clear()
                else:
                    conn.commit()

            except Exception as e:
                conn.rollback()
                result.success_added.clear()
                result.success_updated.clear()
                result.errors.append({
                    "row": "all",
                    "error": str(e)
                })

        except Exception as e:
            logging.error(f"Critical error during import: {str(e)}")
            raise ValueError(str(e))

        finally:
            try:
                wb.close()
            except:
                pass
            conn.close()

        return result 